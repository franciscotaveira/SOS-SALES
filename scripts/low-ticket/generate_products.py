#!/usr/bin/env python3
"""Generate the first sellable artifact for every low-ticket catalog hypothesis.

The catalog is intentionally data-driven: one product manifest feeds the local
deliverables, the Cakto import sheet, and the public product vitrine. Nothing in
this script creates a Cakto account or publishes an offer remotely.
"""

from __future__ import annotations

import json
import re
import shutil
import unicodedata
import zipfile
from datetime import date
from html import escape
from pathlib import Path

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    KeepTogether,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)


def _spreadsheet_dependencies():
    """Load the bundled spreadsheet runtime only when a product needs it."""
    from openpyxl import Workbook
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    return Workbook, ColorScaleRule, Alignment, Border, Font, PatternFill, Side, DataValidation


ROOT = Path(__file__).resolve().parents[2]
CATALOG_PATH = ROOT / "docs" / "CATALOGO_POSSIBILIDADES_LOW_TICKET_2026-09-05.md"
OUT = ROOT / "products" / "low-ticket"
WEB = ROOT / "landing" / "produtos"
WEB_ASSETS = WEB / "assets"
RELEASES = OUT / "releases"
DOWNLOADS = WEB / "downloads"
CAKTO_CHECKOUTS_PATH = OUT / "cakto-checkouts.json"

FAMILY_META = {
    "Mensagens para copiar, adaptar e enviar": {
        "key": "scripts-whatsapp",
        "label": "Scripts para WhatsApp",
        "image": "01-scripts-whatsapp.png",
        "buyer": "vendedor, autônomo ou dono que atende pelo WhatsApp",
        "promise": "responder com clareza, contexto e um próximo passo sem transformar a conversa em robô",
        "prices": [27, 27, 27, 27, 27, 27],
        "metric": "acesso único ao kit de mensagens e rotina",
        "angle": "uma conversa comercial específica",
    },
    "Apresentação comercial pronta para editar": {
        "key": "apresentacao-comercial",
        "label": "Apresentação comercial",
        "image": "02-apresentacao-comercial.png",
        "buyer": "prestador de serviço, pequeno comércio ou profissional independente",
        "promise": "apresentar uma oferta de forma simples para o cliente entender e decidir",
        "prices": [37, 37, 37, 37, 27, 47],
        "metric": "acesso único ao modelo editável e exemplo",
        "angle": "um material comercial que sai pronto para adaptar",
    },
    "Conteúdo e entrada de conversas": {
        "key": "conteudo-campanhas",
        "label": "Conteúdo e campanhas",
        "image": "03-conteudo-campanhas.png",
        "buyer": "dono de negócio, vendedor ou responsável por conteúdo",
        "promise": "transformar uma oferta e dúvidas reais em convites claros para conversar",
        "prices": [27, 47, 37, 47, 67, 47],
        "metric": "acesso único ao roteiro, matriz e modelos de campanha",
        "angle": "uma oferta que precisa começar conversas",
    },
    "Agenda, comparecimento e atendimento de serviços": {
        "key": "agenda-servicos",
        "label": "Agenda e atendimento",
        "image": "04-agenda-servicos.png",
        "buyer": "negócio de serviços com horários marcados",
        "promise": "organizar confirmação, remarcação e orientação sem prometer uma disponibilidade não consultada",
        "prices": [27, 37, 37, 27, 47, 27],
        "metric": "acesso único ao fluxo e às mensagens de atendimento",
        "angle": "uma etapa de agenda que costuma gerar retrabalho",
    },
    "Pós-venda e relacionamento com clientes": {
        "key": "pos-venda",
        "label": "Pós-venda e relacionamento",
        "image": "05-pos-venda.png",
        "buyer": "negócio que já possui clientes e quer cuidar da continuidade",
        "promise": "criar contatos pertinentes depois da compra, com critério, contexto e respeito à preferência do cliente",
        "prices": [37, 37, 47, 47, 37, 47],
        "metric": "acesso único à sequência, matriz e controle de relacionamento",
        "angle": "uma oportunidade de continuidade depois da venda",
    },
    "Planilhas e pequenas ferramentas comerciais": {
        "key": "ferramentas-comerciais",
        "label": "Ferramentas comerciais",
        "image": "06-ferramentas-comerciais.png",
        "buyer": "dono, autônomo ou gestor de pequena equipe",
        "promise": "tirar uma decisão comercial da cabeça e colocar em uma ferramenta simples de usar",
        "prices": [47, 67, 37, 47, 67, 47],
        "metric": "acesso único à planilha, exemplo e instruções",
        "angle": "uma decisão comercial que precisa ficar visível",
    },
    "Treinamento e padrão de atendimento": {
        "key": "treinamento",
        "label": "Treinamento e padrão",
        "image": "07-treinamento.png",
        "buyer": "dono ou líder que precisa orientar vendedores",
        "promise": "dar ao time uma forma observável de praticar, atender e passar a conversa adiante",
        "prices": [67, 67, 47, 67, 67, 47],
        "metric": "acesso único ao treinamento curto, fichas e modelos",
        "angle": "um padrão que o time consegue praticar",
    },
    "Produtos para quem vende implantação de IA": {
        "key": "implantacao-ia",
        "label": "Implantação de IA",
        "image": "08-implantacao-ia.png",
        "buyer": "freelancer, consultor ou pequena agência de automação",
        "promise": "vender e entregar um projeto de IA com escopo, responsabilidade e manutenção explícitos",
        "prices": [97, 27, 67, 67, 47, 97],
        "metric": "acesso único ao template profissional e material de aplicação",
        "angle": "uma etapa profissional de implantação de IA",
    },
}


SCRIPT_PACKS = {
    "01": [
        ("Receber a pergunta de preço", "Oi, [NOME]. O valor de [OFERTA] é [PREÇO] na condição [CONDIÇÃO]. Para eu te orientar sem te empurrar algo errado: você procura isso para [USO/OBJETIVO]?"),
        ("Entender o contexto", "Entendi. O que pesa mais para você agora: [CRITÉRIO A] ou [CRITÉRIO B]?"),
        ("Conectar oferta e necessidade", "Pelo que você contou, [OFERTA] faz sentido quando [SITUAÇÃO]. Ela inclui [INCLUSÃO CONFIRMADA]."),
        ("Abrir o próximo passo", "Se fizer sentido, posso te enviar [PRÓXIMO PASSO REAL: LINK / HORÁRIO / ORÇAMENTO]."),
        ("Quando falta informação", "Não quero adivinhar essa condição. Vou confirmar [DADO] com a equipe e retorno pelo canal combinado."),
        ("Encerrar com respeito", "Se agora não for o momento, tudo bem. Posso encerrar este acompanhamento e você me chama quando quiser retomar."),
    ],
    "02": [
        ("Confirmar recebimento", "Oi, [NOME]. Você conseguiu abrir o orçamento de [OFERTA]?"),
        ("Localizar a dúvida", "Para eu ajudar de forma objetiva: a dúvida ficou em [ESCOPO], [PRAZO], [VALOR] ou outro ponto?"),
        ("Reforçar o escopo", "O orçamento considera [ITENS INCLUÍDOS] e não inclui [EXCLUSÃO]. Se algo mudou, eu reviso o escopo antes de falar em valor."),
        ("Retomar sem pressão", "Estou retomando o orçamento de [OFERTA] porque combinamos falar em [MOMENTO]. Ainda faz sentido avaliar ou prefere pausar?"),
        ("Definir prazo de acompanhamento", "Posso deixar este orçamento em acompanhamento até [DATA REAL]. Depois disso, encerro o contato para não te incomodar."),
        ("Encerrar", "Como não tive retorno, vou encerrar este acompanhamento por agora. Se a necessidade voltar, me escreva e retomamos do ponto certo."),
    ],
    "03": [
        ("Preço", "Entendo comparar preço. Para comparar corretamente, veja [DIFERENCIAL REAL] e [LIMITE REAL]. A condição aprovada é [CONDIÇÃO]."),
        ("Vou pensar", "Claro. O que você quer avaliar antes de decidir: [VALOR], [ESCOPO], [PRAZO] ou [CONFIANÇA]?"),
        ("Achei mais barato", "Pode fazer sentido. Você consegue me dizer se as duas propostas incluem [ITEM ESSENCIAL]? Assim comparamos o escopo, não só o número."),
        ("Preciso falar com alguém", "Perfeito. Posso te enviar um resumo curto com [PONTOS] para facilitar essa conversa?"),
        ("Desconto", "A condição aprovada para [OFERTA] é [CONDIÇÃO]. Não consigo alterar por aqui, mas posso encaminhar uma solicitação de exceção se essa regra existir."),
        ("Sem encaixe", "Pelo que você busca, talvez [OFERTA] não seja o melhor encaixe agora. Prefiro te dizer isso antes de você contratar."),
    ],
    "04": [
        ("Retomar por interesse declarado", "Oi, [NOME]. Você tinha comentado sobre [NECESSIDADE]. Isso ainda está em prioridade ou mudou?"),
        ("Retomar por etapa", "A última etapa foi [ETAPA REAL]. Ficou alguma pendência para decidir o próximo passo?"),
        ("Oferecer ajuda", "Se ajudar, posso resumir em uma mensagem o que está definido e o que ainda precisa de confirmação."),
        ("Perguntar o momento", "Você prefere retomar agora, em [DATA COMBINADA], ou encerrar este acompanhamento?"),
        ("Parar quando não há sinal", "Como não identifiquei uma próxima ação combinada, vou parar por aqui. Você pode chamar quando for oportuno."),
        ("Registrar o motivo", "Para eu organizar corretamente: você pausou por [MOTIVO], escolheu outra opção ou a necessidade deixou de existir?"),
    ],
    "05": [
        ("Resumir a escolha", "Só para confirmar: você escolheu [ITEM/PLANO], com [CONDIÇÃO], e o próximo passo é [AÇÃO]. Está correto?"),
        ("Orientar pagamento", "O pagamento deve ser feito somente pelo [CANAL APROVADO]. Depois da confirmação, [PRÓXIMA ETAPA REAL]."),
        ("Confirmar pedido", "Recebi a confirmação de [STATUS REAL]. Vou registrar seu pedido e [AÇÃO DA EQUIPE]."),
        ("Ajustar um dado", "Antes de finalizar, preciso confirmar [DADO]. Não vou concluir enquanto essa informação estiver pendente."),
        ("Explicar entrega", "A entrega prevista é [ENTREGA CONFIRMADA]. Se surgir uma exceção, a equipe responsável confirma diretamente com você."),
        ("Fechar com registro", "Pedido registrado. Se precisar falar sobre este pedido, envie [REFERÊNCIA] para encontrarmos o contexto sem repetir tudo."),
    ],
    "06": [
        ("Reconhecer o problema", "Entendi o que aconteceu com [SITUAÇÃO]. Vou registrar os fatos sem presumir a causa."),
        ("Pedir o mínimo necessário", "Para encaminhar corretamente, preciso apenas de [DADO MÍNIMO]. Evite enviar dados sensíveis por aqui."),
        ("Sinalizar limite", "Não consigo aprovar [REEMBOLSO/PRAZO/EXCEÇÃO] por esta conversa. Vou encaminhar para [RESPONSÁVEL]."),
        ("Resumir para a equipe", "Vou encaminhar: pedido [X], ocorrido [Y], impacto informado [Z] e retorno esperado [SE EXISTIR]."),
        ("Não prometer", "Ainda não tenho uma solução ou prazo confirmado. Assim que houver uma decisão, [CANAL DE RETORNO]."),
        ("Encerrar o handoff", "Seu caso foi encaminhado para [EQUIPE]. A partir daqui, o responsável continua com este contexto."),
    ],
}


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"[^a-zA-Z0-9]+", "-", normalized.lower()).strip("-")
    return normalized


def parse_catalog() -> list[dict[str, str]]:
    family = None
    records: list[dict[str, str]] = []
    for raw in CATALOG_PATH.read_text(encoding="utf-8").splitlines():
        if raw.startswith("## "):
            family = re.sub(r"^\d+\.\s*", "", raw[3:].strip())
            continue
        match = re.match(r"\|\s*(\d{2})\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|$", raw)
        if not match or family not in FAMILY_META:
            continue
        product_id, title, delivery, mark = match.groups()
        meta = FAMILY_META[family]
        price_index = (int(product_id) - 1) % 6
        records.append(
            {
                "id": product_id,
                "title": title,
                "delivery": delivery,
                "mark": mark,
                "family": family,
                "familyKey": meta["key"],
                "familyLabel": meta["label"],
                "image": meta["image"],
                "buyer": meta["buyer"],
                "promise": meta["promise"],
                "price": meta["prices"][price_index],
                "metric": meta["metric"],
                "angle": meta["angle"],
                "slug": f"p{product_id}-{slugify(title)}",
            }
        )
    if len(records) != 48:
        raise RuntimeError(f"Expected 48 catalog products, found {len(records)}")
    return records


def common_guardrails(record: dict[str, str]) -> list[str]:
    return [
        "Use somente informações que a empresa confirmou.",
        "Troque os campos entre colchetes antes de enviar ou publicar.",
        "Não invente preço, prazo, disponibilidade, estoque, capacidade ou resultado.",
        "Registre uma próxima ação ou encerre o acompanhamento; não insista sem contexto.",
        "Quando houver exceção, encaminhe para uma pessoa responsável.",
    ]


def script_material(record: dict[str, str]) -> list[str]:
    scenarios = SCRIPT_PACKS[record["id"]]
    lines = [
        "## Como usar em 15 minutos",
        "1. Escolha somente as situações que acontecem na sua operação.",
        "2. Substitua os campos `[ENTRE COLCHETES]` com dados aprovados.",
        "3. Envie uma mensagem por vez e aguarde a resposta antes de avançar.",
        "",
        "## Biblioteca de mensagens",
    ]
    for index, (name, message) in enumerate(scenarios, 1):
        lines.extend([f"### {index}. {name}", f"> {message}", "", "**Quando parar:** se não houver resposta, necessidade ou próxima ação confirmada, encerre o acompanhamento.", ""])
    lines.extend(
        [
            "## Ficha de adaptação",
            "| Campo | Preenchimento aprovado |",
            "|---|---|",
            "| Oferta ou serviço | [PREENCHER] |",
            "| Público principal | [PREENCHER] |",
            "| Próximo passo permitido | [PREENCHER] |",
            "| Condição que precisa de confirmação | [PREENCHER] |",
            "| Pessoa que assume exceções | [PREENCHER] |",
            "",
            "## Teste antes de usar",
            "Envie cada mensagem para uma conversa de teste. Verifique se o texto soa humano, se não repete informação já dada e se deixa claro o próximo passo.",
        ]
    )
    return lines


def presentation_material(record: dict[str, str]) -> list[str]:
    templates = {
        "07": ("Catálogo de bolso", "uma página por oferta, com uma escolha simples e um CTA", ["Nome da oferta", "Para quem é", "O que resolve", "O que inclui", "Preço/condição aprovada", "Próximo passo"]),
        "08": ("Orçamento em uma página", "uma proposta que explicita escopo e limites", ["Problema entendido", "Entregas", "Fora do escopo", "Prazo", "Investimento", "Validade e aceite"]),
        "09": ("Três opções", "um comparativo honesto entre essencial, recomendado e completo", ["Critério de escolha", "O que muda em cada opção", "Para quem serve", "Prazo", "Investimento", "Recomendação condicional"]),
        "10": ("Caso de trabalho", "uma prova estruturada sem inventar resultado", ["Contexto autorizado", "Problema", "Intervenção", "Evidência disponível", "Limite do relato", "Próximo passo"]),
        "11": ("Áudio comercial", "um roteiro de 45 a 75 segundos que conversa com a dúvida real", ["Abertura", "Contexto", "Explicação", "Condição", "Convite", "Pausa para resposta"]),
        "12": ("Demonstração em vídeo", "um storyboard curto para mostrar uso e decisão", ["Cena 1: situação", "Cena 2: ação", "Cena 3: evidência", "Cena 4: limite", "Cena 5: próximo passo", "Legenda e CTA"]),
    }
    name, outcome, fields = templates[record["id"]]
    lines = [
        "## Resultado do material",
        f"Você vai sair com {outcome}.",
        "",
        "## Modelo preenchível",
        "| Campo | Sua versão |",
        "|---|---|",
    ]
    for field in fields:
        lines.append(f"| {field} | [PREENCHER] |")
    lines.extend(
        [
            "",
            "## Exemplo fictício",
            f"**{name}:** uma empresa de manutenção residencial apresenta [SERVIÇO] para [PÚBLICO], inclui [ITEM CONFIRMADO], informa [PRAZO APROVADO] e convida a pessoa a [PRÓXIMO PASSO]. O exemplo é ilustrativo; substitua todos os dados antes de publicar.",
            "",
            "## Revisão antes de enviar",
            "- O cliente entende o que recebe em menos de um minuto?",
            "- O preço e a condição foram conferidos?",
            "- Há alguma promessa sem fonte ou autorização?",
            "- Existe uma ação clara para responder?",
        ]
    )
    return lines


def content_material(record: dict[str, str]) -> list[str]:
    content_map = {
        "13": ("Status do WhatsApp", ["Demonstre uma situação", "Mostre uma dúvida frequente", "Explique uma escolha", "Convide para responder", "Resuma uma condição", "Abra uma conversa", "Encerre com orientação"]),
        "14": ("Campanha de uma oferta", ["Preparação", "Abertura", "Prova ou demonstração", "Objeção", "Convite", "Lembrete pertinente", "Encerramento real"]),
        "15": ("Dúvidas dos clientes", ["Colete a pergunta", "Separe fato de hipótese", "Escolha o formato", "Escreva a resposta", "Inclua o próximo passo", "Publique e registre", "Revise o que voltou"]),
        "16": ("Vários criativos", ["Ângulo de problema", "Ângulo de resultado", "Ângulo de escolha", "Ângulo de demonstração", "Ângulo de objeção", "Ângulo de identidade", "Ângulo de teste"]),
        "17": ("Anúncio à conversa", ["Promessa", "Chamada", "Intenção esperada", "Primeira resposta", "Dado mínimo", "Próximo passo", "Sinal de perda"]),
        "18": ("Parceria local", ["Escolha do parceiro", "Contexto da indicação", "Mensagem individual", "Oferta de colaboração", "Registro do contato", "Retorno", "Encerramento"]),
    }
    title, days = content_map[record["id"]]
    lines = [
        "## Roteiro de aplicação",
        f"Use a sequência **{title}** como uma semana de teste. Uma peça precisa ter uma intenção observável: informar, esclarecer ou iniciar uma conversa.",
        "",
        "| Dia | Intenção | Peça ou ação | Sinal observado |",
        "|---:|---|---|---|",
    ]
    for day, label in enumerate(days, 1):
        lines.append(f"| {day} | {label} | [PREENCHER] | [REGISTRAR] |")
    lines.extend(
        [
            "",
            "## Briefing de uma peça",
            "- Oferta real: [PREENCHER]",
            "- Público que já demonstrou essa necessidade: [PREENCHER]",
            "- Frase que o cliente costuma usar: [PREENCHER]",
            "- O que a pessoa deve fazer depois: [PREENCHER]",
            "- O que não pode ser prometido: [PREENCHER]",
            "",
            "## Leitura do teste",
            "Registre respostas, perguntas e conversas iniciadas. Não atribua uma venda ao conteúdo sem confirmar a origem com o cliente ou no seu registro comercial.",
        ]
    )
    return lines


def agenda_material(record: dict[str, str]) -> list[str]:
    agenda_map = {
        "19": ("Confirmar", ["convite", "data e horário", "local ou canal", "preparo", "resposta esperada"]),
        "20": ("Remarcar ou cancelar", ["pedido recebido", "política existente", "alternativas", "confirmação", "registro do motivo"]),
        "21": ("Lista de espera", ["preferência", "critério de convite", "disponibilidade confirmada", "resposta", "remoção da lista"]),
        "22": ("Após uma falta", ["reconhecer ausência", "perguntar se deseja retomar", "oferecer somente opções reais", "registrar", "encerrar"]),
        "23": ("Consulta à proposta", ["objetivo", "dados mínimos", "resumo", "escopo", "próximo passo"]),
        "24": ("Primeira visita", ["boas-vindas", "localização", "documentos", "preparo", "confirmação"]),
    }
    label, steps = agenda_map[record["id"]]
    lines = [
        "## Fluxo de atendimento",
        f"O fluxo **{label}** só confirma o que a empresa consegue consultar ou cumprir.",
        "",
        "| Etapa | Mensagem ou decisão | Fonte de confirmação |",
        "|---:|---|---|",
    ]
    for index, step in enumerate(steps, 1):
        lines.append(f"| {index} | [PREENCHER: {step}] | [SISTEMA / RESPONSÁVEL] |")
    lines.extend(
        [
            "",
            "## Mensagens base",
            "> Oi, [NOME]. Para [OBJETIVO], posso registrar [DADO] e confirmar [CONDIÇÃO REAL] com a equipe.",
            "> A opção [DATA/HORÁRIO] foi confirmada por [FONTE]. Se precisar alterar, responda até [REGRA APROVADA].",
            "> Ainda não consigo confirmar [DISPONIBILIDADE/PRAZO]. Vou encaminhar para [RESPONSÁVEL] antes de te responder.",
            "",
            "## Limites",
            "Não invente agenda, vaga, estoque, prazo, preparo ou política. Se a informação não estiver integrada e testada, registre a preferência e faça o handoff.",
        ]
    )
    return lines


def post_sale_material(record: dict[str, str]) -> list[str]:
    post_map = {
        "25": ("Acompanhamento", ["recebimento", "primeiro uso", "dúvida", "resultado percebido", "ajuda"]),
        "26": ("Avaliação e depoimento", ["momento certo", "pedido honesto", "permissão", "edição mínima", "registro"]),
        "27": ("Indicação", ["sinal de satisfação", "pedido contextual", "condição", "origem", "agradecimento"]),
        "28": ("Recompra", ["ocasião", "elegibilidade", "exclusão", "oferta pertinente", "resultado"]),
        "29": ("Renovação", ["data", "uso", "condição", "decisão", "registro"]),
        "30": ("Cliente inativo", ["motivo", "contexto", "reconexão", "opção de retorno", "parada"]),
    }
    label, steps = post_map[record["id"]]
    lines = [
        "## Sequência de relacionamento",
        f"A sequência **{label}** precisa respeitar o momento e a preferência do cliente. Use apenas contatos que tenham uma razão clara.",
        "",
        "| Momento | Intenção | Mensagem ou ação | Critério de parada |",
        "|---:|---|---|---|",
    ]
    for index, step in enumerate(steps, 1):
        lines.append(f"| {index} | {step} | [PREENCHER] | [PREENCHER] |")
    lines.extend(
        [
            "",
            "## Mensagens de referência",
            "> Oi, [NOME]. Como foi [MOMENTO REAL] depois de [COMPRA/SERVIÇO]? Se algo ficou pendente, posso registrar para a equipe.",
            "> Se você se sentir confortável, pode contar em uma frase como foi sua experiência. Só usaremos o relato com sua autorização.",
            "> Tenho uma condição de retorno para [OCASIÃO REAL]. Quer que eu explique ou prefere não receber esse tipo de mensagem?",
            "",
            "## Governança",
            "Remova contatos que não devem receber a mensagem, respeite pedidos de pausa e registre quando uma pessoa pediu para não ser contatada.",
        ]
    )
    return lines


def tools_material(record: dict[str, str]) -> list[str]:
    tool_map = {
        "31": ("Follow-up", ["ID", "Nome", "Contexto", "Responsável", "Próxima ação", "Prazo", "Status", "Motivo de parada"]),
        "32": ("Venda rastreável", ["Data", "Origem", "Campanha", "Oportunidade", "Status", "Venda", "Valor", "Motivo de perda"]),
        "33": ("Desconto", ["Preço cheio", "Custo variável", "Desconto", "Preço final", "Contribuição", "Margem"]),
        "34": ("Combo", ["Item", "Quantidade", "Preço unitário", "Custo", "Preço do combo", "Contribuição"]),
        "35": ("Capacidade e metas", ["Pessoas", "Horas", "Atendimentos", "Valor médio", "Conversão", "Cenário"]),
        "36": ("Painel semanal", ["Oportunidades", "Vendas", "Perdas", "Origem", "Próximas ações", "Decisões"]),
    }
    label, columns = tool_map[record["id"]]
    lines = [
        "## Como configurar",
        f"A ferramenta **{label}** começa com um exemplo fictício. Duplique a aba de exemplo, apague os dados de teste e só então registre dados reais.",
        "",
        "### Colunas mínimas",
        "| Coluna | O que registrar |",
        "|---|---|",
    ]
    for column in columns:
        lines.append(f"| {column} | [PREENCHER] |")
    lines.extend(
        [
            "",
            "### Fórmulas de referência",
            "- Preço final = preço cheio × (1 − desconto).",
            "- Contribuição estimada = preço final − custos variáveis informados.",
            "- Conversão observada = vendas ÷ oportunidades elegíveis, quando os dois números usam o mesmo período.",
            "",
            "## Rotina de 20 minutos",
            "1. Atualize somente os registros novos.",
            "2. Marque valores desconhecidos como `desconhecido`; não preencha por suposição.",
            "3. Escolha até três decisões para a semana.",
            "4. Registre quem ficou responsável por cada ação.",
            "",
            "## Limite de interpretação",
            "A planilha organiza evidência. Ela não prova causalidade, garante margem, substitui uma integração ou promete aumento de vendas.",
        ]
    )
    return lines


def training_material(record: dict[str, str]) -> list[str]:
    training_map = {
        "37": ("Primeira semana", ["entender oferta", "observar conversa", "praticar pergunta", "praticar próximo passo", "revisar" ]),
        "38": ("Simulações", ["briefing", "cliente fictício", "rodada", "observação", "nova rodada"]),
        "39": ("Revisão do gestor", ["clareza", "contexto", "informação correta", "próximo passo", "feedback"]),
        "40": ("Passagem de atendimento", ["resumo", "pendência", "responsável", "próxima ação", "confirmação"]),
        "41": ("Manual", ["responsabilidades", "tom", "fontes", "limites", "encaminhamentos"]),
        "42": ("WhatsApp Business", ["perfil", "catálogo", "etiquetas", "respostas rápidas", "revisão"]),
    }
    label, steps = training_map[record["id"]]
    lines = [
        "## Plano de aplicação",
        f"O treinamento **{label}** funciona em uma sessão curta, seguida de observação. O objetivo é criar um padrão observável, não decorar frases.",
        "",
        "| Bloco | Atividade | Evidência de conclusão |",
        "|---:|---|---|",
    ]
    for index, step in enumerate(steps, 1):
        lines.append(f"| {index} | {step} | [O QUE OBSERVAR] |")
    lines.extend(
        [
            "",
            "## Rubrica de observação",
            "Dê uma nota de 0 a 2 e escreva uma evidência: 0 = não apareceu; 1 = apareceu com ajuda; 2 = apareceu de forma autônoma.",
            "",
            "| Critério | Nota | Evidência |",
            "|---|---:|---|",
            "| Entendeu o contexto antes de responder |  |  |",
            "| Informou somente o que estava aprovado |  |  |",
            "| Fez uma pergunta útil |  |  |",
            "| Deixou o próximo passo claro |  |  |",
            "| Soube transferir ou encerrar |  |  |",
            "",
            "## Feedback",
            "Comece pelo comportamento observável, mostre um exemplo e peça uma nova tentativa. Evite avaliação de personalidade ou cobrança por uma resposta decorada.",
        ]
    )
    return lines


def professional_material(record: dict[str, str]) -> list[str]:
    professional_map = {
        "43": ("Proposta de automação", ["contexto", "objetivo", "escopo", "exclusões", "responsabilidades", "marcos", "aceite"]),
        "44": ("Calculadora de setup", ["setup", "recorrência", "consumo", "suporte", "contingência", "contribuição"]),
        "45": ("Briefing de projeto", ["processo atual", "usuário", "fontes", "integrações", "risco", "viabilidade"]),
        "46": ("Onboarding de implantação", ["dependências", "responsáveis", "acessos", "validação", "treinamento", "passagem"]),
        "47": ("Demonstração", ["cenário", "problema", "ação", "evidência", "limite", "próximo passo"]),
        "48": ("Entrega e manutenção", ["versão", "alterações", "aceite", "suporte", "incidentes", "próxima revisão"]),
    }
    label, fields = professional_map[record["id"]]
    lines = [
        "## Artefato profissional",
        f"O template **{label}** deve ser preenchido junto do cliente e aprovado antes de iniciar a próxima etapa.",
        "",
        "| Campo | Registro do projeto | Responsável |",
        "|---|---|---|",
    ]
    for field in fields:
        lines.append(f"| {field} | [PREENCHER] | [PREENCHER] |")
    lines.extend(
        [
            "",
            "## Perguntas de controle",
            "- O cliente entende o que está incluído e o que está fora do escopo?",
            "- Há uma fonte aprovada para cada informação que o agente usa?",
            "- Quem assume decisões, exceções e incidentes?",
            "- Como o aceite será evidenciado?",
            "- Qual é a próxima revisão e o que a manutenção inclui?",
            "",
            "## Limites comerciais",
            "Não prometa conformidade, disponibilidade, comportamento perfeito, redução de custo ou aumento de vendas sem evidência e escopo específico. Não solicite senhas no briefing; registre dependências e use o processo seguro do projeto.",
        ]
    )
    return lines


TOOL_FILE_NAMES = {
    "15": "planilha-perguntas-conteudo.xlsx",
    "18": "quadro-parcerias.xlsx",
    "19": "fluxo-confirmacao.xlsx",
    "20": "fluxo-remarcacao.xlsx",
    "21": "lista-espera.xlsx",
    "25": "rotina-pos-venda.xlsx",
    "27": "controle-indicacoes.xlsx",
    "28": "matriz-recompra.xlsx",
    "29": "calendario-renovacao.xlsx",
    "30": "reativacao-clientes.xlsx",
    "31": "follow-up.xlsx",
    "32": "venda-rastreavel.xlsx",
    "33": "calculadora-desconto.xlsx",
    "34": "calculadora-combos.xlsx",
    "35": "planejador-capacidade-metas.xlsx",
    "36": "painel-comercial-semanal.xlsx",
    "44": "calculadora-setup-mensalidade.xlsx",
}


def build_tool_file(record: dict[str, str], path: Path) -> str | None:
    """Create a small, editable workbook for products whose promise is a tool.

    The first sheet is always an instruction sheet and example rows are clearly
    labelled. Formulas are intentionally transparent so the buyer can audit
    the assumptions instead of receiving a black-box calculator.
    """
    filename = TOOL_FILE_NAMES.get(record["id"])
    if not filename:
        return None
    Workbook, ColorScaleRule, Alignment, Border, Font, PatternFill, Side, DataValidation = _spreadsheet_dependencies()

    wb = Workbook()
    readme = wb.active
    readme.title = "LEIA-ME"
    dark = "0B0F17"
    green = "0A8A65"
    lime = "B7F66B"
    paper = "F5F8F6"
    line = "C8D9D5"
    heading_fill = PatternFill("solid", fgColor=dark)
    input_fill = PatternFill("solid", fgColor="FFF3C4")
    thin = Side(style="thin", color=line)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def title_sheet(ws, title, subtitle):
        ws.sheet_view.showGridLines = False
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        ws["A1"].fill = heading_fill
        ws.merge_cells("A1:H1")
        ws["A2"] = subtitle
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells("A2:H2")
        ws.row_dimensions[2].height = 34
        ws.freeze_panes = "A4"

    def headers(ws, row, labels):
        for col, label in enumerate(labels, 1):
            cell = ws.cell(row=row, column=col, value=label)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = heading_fill
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    def table_range(ws, min_row, max_row, max_col):
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    readme["A1"] = record["title"]
    readme["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    readme["A1"].fill = heading_fill
    readme.merge_cells("A1:F1")
    readme["A3"] = "Comece aqui"
    readme["A3"].font = Font(bold=True, size=12, color=green)
    readme["A4"] = "1. Leia a aba de instruções do modelo."
    readme["A5"] = "2. Apague ou preserve apenas a linha marcada como EXEMPLO, conforme sua necessidade."
    readme["A6"] = "3. Substitua os campos amarelos por dados reais e confirmados."
    readme["A7"] = "4. Registre a data da última revisão e a pessoa responsável."
    readme["A9"] = "Limites"
    readme["A9"].font = Font(bold=True, size=12, color=green)
    readme["A10"] = "Esta ferramenta organiza evidência e premissas. Ela não garante margem, agenda, conversão, disponibilidade ou resultado."
    readme["A10"].alignment = Alignment(wrap_text=True, vertical="top")
    readme.merge_cells("A10:F11")
    readme["A13"] = "Campos de entrada estão em amarelo; fórmulas ficam visíveis para auditoria."
    readme["A13"].font = Font(italic=True, color="5F706C")
    for col, width in {"A": 28, "B": 20, "C": 20, "D": 20, "E": 20, "F": 20}.items():
        readme.column_dimensions[col].width = width
    readme.sheet_view.showGridLines = False

    pid = record["id"]
    if pid in {"31", "32", "18", "19", "20", "21", "25", "27", "28", "29", "30"}:
        ws = wb.create_sheet("Registros")
        if pid == "31":
            labels = ["ID", "Nome", "Contexto", "Responsável", "Próxima ação", "Prazo", "Status", "Motivo de parada"]
            example = ["EXEMPLO", "Pessoa exemplo", "Pediu orçamento de [OFERTA]", "[RESPONSÁVEL]", "Confirmar escopo", "2026-09-12", "Em andamento", ""]
            subtitle = "Use uma linha por oportunidade. Uma próxima ação sem prazo não é uma próxima ação operável."
        elif pid == "32":
            labels = ["Data", "Origem", "Campanha", "Oportunidade", "Status", "Venda", "Valor", "Motivo de perda"]
            example = ["2026-09-05", "WhatsApp", "EXEMPLO", "Oferta exemplo", "Ganha", "SIM", 97, ""]
            subtitle = "Preserve origem desconhecida quando não houver evidência; não atribua venda por suposição."
        elif pid == "18":
            labels = ["Parceiro", "Contexto da indicação", "Contato", "Oferta de colaboração", "Responsável", "Próximo passo", "Data", "Status"]
            example = ["Parceiro exemplo", "Atende o mesmo público", "[CONTATO]", "Troca de indicação", "[RESPONSÁVEL]", "Enviar apresentação", "2026-09-12", "Em conversa"]
            subtitle = "Registre o contexto para que a indicação seja pertinente e autorizada."
        elif pid in {"19", "20", "21"}:
            labels = ["Nome", "Contato", "Preferência/data", "Condição confirmada", "Responsável", "Próxima ação", "Status", "Observação"]
            example = ["Pessoa exemplo", "[CONTATO]", "[DATA/HORÁRIO]", "[FONTE]", "[RESPONSÁVEL]", "Confirmar", "Pendente", "EXEMPLO — substituir"]
            subtitle = "Nunca transforme preferência em disponibilidade. Confirme a condição na fonte indicada."
        else:
            labels = ["Nome", "Contato", "Ocasião", "Última interação", "Próxima ação", "Responsável", "Status", "Motivo de parada"]
            example = ["Pessoa exemplo", "[CONTATO]", "[MOMENTO REAL]", "2026-09-05", "Enviar mensagem pertinente", "[RESPONSÁVEL]", "Pendente", ""]
            subtitle = "Use contexto e pertinência para decidir se um novo contato faz sentido."
        title_sheet(ws, record["title"], subtitle)
        headers(ws, 3, labels)
        for col, value in enumerate(example, 1):
            ws.cell(4, col, value=value).fill = input_fill
        for row in range(5, 34):
            for col in range(1, len(labels) + 1):
                ws.cell(row, col).fill = input_fill
        table_range(ws, 3, 33, len(labels))
        ws.auto_filter.ref = f"A3:{chr(64 + len(labels))}33"
        for col in range(1, len(labels) + 1):
            ws.column_dimensions[chr(64 + col)].width = 19
        if pid == "31":
            dv = DataValidation(type="list", formula1='"Pendente,Em andamento,Concluído,Encerrado"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add("G4:G33")
        elif pid == "32":
            dv = DataValidation(type="list", formula1='"SIM,NÃO"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add("F4:F33")
            ws["G4"].number_format = 'R$ #,##0.00'
            for row in range(5, 34):
                ws.cell(row, 7).number_format = 'R$ #,##0.00'
    elif pid in {"33", "34", "35", "36", "44"}:
        ws = wb.create_sheet("Calculadora")
        if pid == "33":
            title_sheet(ws, record["title"], "Preencha os valores amarelos; o resultado mostra o efeito do desconto informado.")
            labels = [("Preço cheio", 100), ("Custo variável", 35), ("Desconto (%)", 10)]
            headers(ws, 3, ["Entrada", "Valor"])
            for row, (label, value) in enumerate(labels, 4):
                ws.cell(row, 1, label); ws.cell(row, 2, value); ws.cell(row, 2).fill = input_fill
            headers(ws, 9, ["Saída", "Fórmula", "Resultado"])
            outputs = [("Preço final", "Preço cheio × (1 − desconto)", "=B4*(1-B6/100)"), ("Contribuição", "Preço final − custo variável", "=C10-B5"), ("Margem sobre preço", "Contribuição ÷ preço final", "=IFERROR(C11/C10,0)")]
            for row, (label, formula, value) in enumerate(outputs, 10):
                ws.cell(row, 1, label); ws.cell(row, 2, formula); ws.cell(row, 3, value)
            ws["B4"].number_format = ws["B5"].number_format = ws["C10"].number_format = ws["C11"].number_format = 'R$ #,##0.00'
            ws["B6"].number_format = '0.0'
            ws["C12"].number_format = '0.0%'
            for row in range(10, 13): table_range(ws, row, row, 3)
        elif pid == "34":
            title_sheet(ws, record["title"], "Compare itens avulsos e pacote com quantidade, preço, custo e contribuição estimada.")
            headers(ws, 3, ["Item", "Quantidade", "Preço unitário", "Custo unitário", "Receita", "Custo", "Contribuição"])
            sample = [["Item exemplo", 1, 60, 20], ["Segundo item", 1, 40, 12]]
            for row, values in enumerate(sample, 4):
                for col, value in enumerate(values, 1): ws.cell(row, col, value=value).fill = input_fill
                ws.cell(row, 5, f"=B{row}*C{row}"); ws.cell(row, 6, f"=B{row}*D{row}"); ws.cell(row, 7, f"=E{row}-F{row}")
            for row in range(6, 14):
                for col in range(1, 5): ws.cell(row, col).fill = input_fill
                ws.cell(row, 5, f"=B{row}*C{row}"); ws.cell(row, 6, f"=B{row}*D{row}"); ws.cell(row, 7, f"=E{row}-F{row}")
            headers(ws, 16, ["Resumo", "Valor"])
            ws["A17"] = "Receita total"; ws["B17"] = "=SUM(E4:E13)"
            ws["A18"] = "Contribuição total"; ws["B18"] = "=SUM(G4:G13)"
            for cell in ("B17", "B18"): ws[cell].number_format = 'R$ #,##0.00'
            table_range(ws, 3, 13, 7); table_range(ws, 16, 18, 2)
        elif pid == "35":
            title_sheet(ws, record["title"], "Simule capacidade e metas com premissas que o usuário consegue conferir.")
            headers(ws, 3, ["Premissa", "Valor"])
            entries = [("Pessoas", 1), ("Horas úteis/dia", 6), ("Dias úteis/mês", 20), ("Atendimentos/hora", 3), ("Valor médio", 150), ("Conversão (%)", 20)]
            for row, (label, value) in enumerate(entries, 4):
                ws.cell(row, 1, label); ws.cell(row, 2, value); ws.cell(row, 2).fill = input_fill
            headers(ws, 12, ["Indicador", "Fórmula", "Resultado"])
            outputs = [("Capacidade mensal", "pessoas × horas × dias × atendimentos/hora", "=B4*B5*B6*B7"), ("Receita potencial", "capacidade × valor médio × conversão", "=C13*B8*B9/100"), ("Vendas estimadas", "capacidade × conversão", "=C13*B9/100")]
            for row, (label, formula, value) in enumerate(outputs, 13): ws.cell(row,1,label); ws.cell(row,2,formula); ws.cell(row,3,value)
            ws["B8"].number_format = ws["C14"].number_format = 'R$ #,##0.00'; ws["B9"].number_format = '0.0'
            table_range(ws, 3, 9, 2); table_range(ws, 12, 15, 3)
        elif pid == "36":
            title_sheet(ws, record["title"], "Atualize os números da mesma semana e escolha poucas decisões executáveis.")
            headers(ws, 3, ["Indicador", "Semana atual", "Semana anterior", "Observação"])
            for row, label in enumerate(["Oportunidades", "Vendas", "Perdas", "Valor vendido", "Próximas ações", "Decisões"], 4):
                ws.cell(row, 1, label); ws.cell(row, 2).fill = input_fill; ws.cell(row, 3).fill = input_fill; ws.cell(row, 4).fill = input_fill
            ws["A12"] = "Conversão observada"; ws["B12"] = "=IFERROR(B5/B4,0)"; ws["C12"] = "=IFERROR(C5/C4,0)"
            ws["B12"].number_format = ws["C12"].number_format = '0.0%'
            table_range(ws, 3, 12, 4)
        else:
            title_sheet(ws, record["title"], "Compare setup, recorrência, consumo e suporte antes de apresentar uma proposta.")
            headers(ws, 3, ["Premissa", "Valor"])
            entries = [("Setup desejado", 1500), ("Mensalidade desejada", 400), ("Consumo estimado/mês", 120), ("Suporte estimado/mês", 100), ("Contingência (%)", 10)]
            for row, (label, value) in enumerate(entries, 4):
                ws.cell(row, 1, label); ws.cell(row, 2, value); ws.cell(row, 2).fill = input_fill
            headers(ws, 11, ["Indicador", "Fórmula", "Resultado"])
            outputs = [("Custo mensal estimado", "consumo + suporte", "=B6+B7"), ("Receita mensal após custos", "mensalidade − custo mensal", "=B5-C12"), ("Setup líquido de contingência", "setup × (1 − contingência)", "=B4*(1-B8/100)")]
            for row, (label, formula, value) in enumerate(outputs, 12): ws.cell(row,1,label); ws.cell(row,2,formula); ws.cell(row,3,value)
            for cell in ("B4", "B5", "B6", "B7", "C12", "C13", "C14"): ws[cell].number_format = 'R$ #,##0.00'
            ws["B8"].number_format = '0.0'
            table_range(ws, 3, 8, 2); table_range(ws, 11, 14, 3)
        for col in range(1, 9): ws.column_dimensions[chr(64 + col)].width = 21
        ws.conditional_formatting.add("C10:C12", ColorScaleRule(start_type="min", start_color="F8D7DA", mid_type="percentile", mid_value=50, mid_color="FFF3C4", end_type="max", end_color="B7F66B"))

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    output_path = path / filename
    wb.save(output_path)
    return filename


def family_body(record: dict[str, str]) -> list[str]:
    key = record["familyKey"]
    if key == "scripts-whatsapp":
        return script_material(record)
    if key == "apresentacao-comercial":
        return presentation_material(record)
    if key == "conteudo-campanhas":
        return content_material(record)
    if key == "agenda-servicos":
        return agenda_material(record)
    if key == "pos-venda":
        return post_sale_material(record)
    if key == "ferramentas-comerciais":
        return tools_material(record)
    if key == "treinamento":
        return training_material(record)
    return professional_material(record)


def material_markdown(record: dict[str, str]) -> str:
    lines = [
        f"# {record['title']}",
        "",
        f"**Família:** {record['familyLabel']}  ",
        f"**Para:** {record['buyer']}  ",
        f"**Preço-teste sugerido:** R$ {record['price']},00  ",
        "**Versão:** v1.0 · material de aplicação",
        "",
        "## O que este produto resolve",
        f"{record['promise'].capitalize()}.",
        "",
        "## Entrega incluída",
        f"{record['delivery']}.",
        "",
        "## Comece pelo resultado",
        "Ao terminar, você deve conseguir apontar um artefato preenchido e uma próxima ação observável. O material não inclui implantação, operação ilimitada, tráfego ou consumo de ferramentas.",
        "",
    ]
    lines.extend(family_body(record))
    lines.extend(
        [
            "",
            "## Checklist de segurança e qualidade",
            *[f"- [ ] {item}" for item in common_guardrails(record)],
            "",
            "## Próximo passo",
            "Se este material revelar uma necessidade de operação recorrente, avalie o SOS Vendas. Se revelar uma lacuna na configuração de IA, compare com o EKO antes de comprar outro material.",
        ]
    )
    return "\n".join(lines) + "\n"


def sales_page_markdown(record: dict[str, str]) -> str:
    return "\n".join(
        [
            f"# {record['title']}",
            "",
            f"> {record['promise'].capitalize()}.",
            "",
            f"**Para quem é:** {record['buyer']}.",
            "",
            "## Você recebe",
            f"- {record['delivery']}",
            "- Material de aplicação com campos editáveis",
            "- Exemplo fictício para entender o preenchimento",
            "- Checklist de revisão antes de usar",
            "",
            "## O que não está incluído",
            "Instalação, gestão de campanhas, consultoria ilimitada, garantia de resultado ou informação que a empresa ainda não confirmou.",
            "",
            f"## Investimento de teste: R$ {record['price']},00",
            "Pagamento único. A oferta e as condições finais precisam ser conferidas no checkout Cakto antes de publicar.",
            "",
            "## Perguntas frequentes",
            "**Preciso ser técnico?** Não. O material começa pelo seu processo e mostra onde preencher, testar ou pedir confirmação.",
            "",
            "**Posso usar com qualquer empresa?** O método é adaptável, mas os exemplos e regras precisam ser substituídos pelos dados reais da sua operação.",
            "",
            "**Isso substitui o SOS Vendas ou o EKO?** Não. Este é um material pontual para uma tarefa específica; o CRM atende a operação recorrente e o EKO organiza a configuração comercial da IA.",
            "",
            "## CTA",
            "Quero receber o material e aplicar a primeira versão.",
            "",
            "## Ângulos de divulgação",
            f"1. Dor: resolva {record['angle']} sem começar de uma página em branco.",
            f"2. Resultado: saia com um artefato preenchido para usar na próxima conversa.",
            f"3. Identidade: feito para {record['buyer']}, com campos e limites claros.",
            "",
        ]
    ) + "\n"


def readme_markdown(record: dict[str, str]) -> str:
    tool_file = record.get("toolFile")
    lines = [
            f"# Pacote — {record['title']}",
            "",
            "Arquivos do pacote:",
            "- `material.pdf`: versão pronta para entrega.",
            "- `material.md`: fonte editável.",
            "- `pagina-venda.md`: copy de checkout e divulgação.",
    ]
    if tool_file:
        lines.append(f"- `{tool_file}`: arquivo editável com abas de instrução, exemplo e uso.")
    lines.extend(
        [
            "- A imagem da família do produto aparece na vitrine e na página de venda.",
            "",
            f"Preço-teste sugerido: R$ {record['price']},00.",
            "",
            "Antes de publicar, conferir a oferta no painel Cakto, o conteúdo efetivamente entregue, o e-mail de suporte e o link de checkout.",
            "",
        ]
    )
    return "\n".join(lines)


def markdown_to_story(markdown: str):
    font_path = "/System/Library/Fonts/Supplemental/Arial.ttf"
    bold_path = "/System/Library/Fonts/Supplemental/Arial Bold.ttf"
    if Path(font_path).exists() and "SOSArial" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("SOSArial", font_path))
        pdfmetrics.registerFont(TTFont("SOSArial-Bold", bold_path))
    body_font = "SOSArial" if "SOSArial" in pdfmetrics.getRegisteredFontNames() else "Helvetica"
    bold_font = "SOSArial-Bold" if "SOSArial-Bold" in pdfmetrics.getRegisteredFontNames() else "Helvetica-Bold"
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="SOSBody", parent=styles["BodyText"], fontName=body_font, fontSize=9.2, leading=12.4, textColor=colors.HexColor("#25352f"), spaceAfter=4))
    styles.add(ParagraphStyle(name="SOSH1", parent=styles["Heading1"], fontName=bold_font, fontSize=21, leading=24, textColor=colors.HexColor("#0b0f17"), spaceAfter=10))
    styles.add(ParagraphStyle(name="SOSH2", parent=styles["Heading2"], fontName=bold_font, fontSize=13, leading=16, textColor=colors.HexColor("#0a8a65"), spaceBefore=10, spaceAfter=5))
    styles.add(ParagraphStyle(name="SOSH3", parent=styles["Heading3"], fontName=bold_font, fontSize=10.5, leading=13, textColor=colors.HexColor("#0b0f17"), spaceBefore=7, spaceAfter=3))
    styles.add(ParagraphStyle(name="SOSQuote", parent=styles["SOSBody"], leftIndent=12, borderPadding=6, borderColor=colors.HexColor("#b7f66b"), borderWidth=0.5, borderLeft=True, backColor=colors.HexColor("#f0f7ec"), spaceBefore=3, spaceAfter=7))
    styles.add(ParagraphStyle(name="SOSMeta", parent=styles["SOSBody"], fontSize=8.2, leading=10.5, textColor=colors.HexColor("#5f706c")))
    story = []
    table_buffer: list[str] = []

    def flush_table():
        nonlocal table_buffer
        if not table_buffer:
            return
        rows = []
        for row in table_buffer:
            if set(row.replace("|", "").replace("-", "").replace(":", "").strip()) == set():
                continue
            cells = [cell.strip() for cell in row.strip().strip("|").split("|")]
            if cells and not all(re.fullmatch(r":?-+:?", cell or "-") for cell in cells):
                rows.append([Paragraph(escape(cell).replace("**", ""), styles["SOSBody"]) for cell in cells])
        if rows:
            table = Table(rows, repeatRows=1, hAlign="LEFT", colWidths=None)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0b0f17")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), bold_font),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#c8d9d5")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.extend([Spacer(1, 3), table, Spacer(1, 5)])
        table_buffer = []

    for raw in markdown.splitlines():
        line = raw.strip()
        if line.startswith("|"):
            table_buffer.append(line)
            continue
        flush_table()
        if not line:
            story.append(Spacer(1, 3))
        elif line.startswith("# "):
            story.append(Paragraph(escape(line[2:]), styles["SOSH1"]))
        elif line.startswith("## "):
            story.append(Paragraph(escape(line[3:]), styles["SOSH2"]))
        elif line.startswith("### "):
            story.append(Paragraph(escape(line[4:]), styles["SOSH3"]))
        elif line.startswith("> "):
            story.append(Paragraph(escape(line[2:]), styles["SOSQuote"]))
        elif line.startswith("- [ ] "):
            story.append(Paragraph("☐ " + escape(line[6:]), styles["SOSBody"]))
        elif line.startswith("- "):
            story.append(Paragraph("• " + escape(line[2:]), styles["SOSBody"]))
        elif re.match(r"^\d+\. ", line):
            story.append(Paragraph(escape(line), styles["SOSBody"]))
        else:
            text = escape(line)
            text = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", text)
            text = text.replace("`", "")
            story.append(Paragraph(text, styles["SOSBody"]))
    flush_table()
    return story


def build_pdf(record: dict[str, str], material: str, path: Path) -> None:
    class ProductDocTemplate(BaseDocTemplate):
        def __init__(self, filename, **kwargs):
            super().__init__(filename, **kwargs)
            frame = Frame(self.leftMargin, self.bottomMargin, self.width, self.height, id="normal")
            self.addPageTemplates([PageTemplate(id="product", frames=frame, onPage=self.draw_page)])

        def draw_page(self, canvas, doc):
            canvas.saveState()
            canvas.setFillColor(colors.HexColor("#0b0f17"))
            canvas.rect(0, A4[1] - 15 * mm, A4[0], 15 * mm, fill=1, stroke=0)
            canvas.setFillColor(colors.HexColor("#b7f66b"))
            canvas.setFont("Helvetica-Bold", 8)
            canvas.drawString(18 * mm, A4[1] - 10 * mm, "SOS VENDAS · MATERIAL DE APLICAÇÃO")
            canvas.setFillColor(colors.HexColor("#78918a"))
            canvas.setFont("Helvetica", 7.5)
            canvas.drawRightString(A4[0] - 18 * mm, 11 * mm, f"{record['id']} · v1.0")
            canvas.drawString(18 * mm, 11 * mm, "Use com dados aprovados pela sua empresa")
            canvas.restoreState()

    doc = ProductDocTemplate(str(path), pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm, topMargin=23 * mm, bottomMargin=18 * mm, title=record["title"], author="SOS Vendas")
    doc.build(markdown_to_story(material))


def write_catalog_js(records: list[dict[str, str]]) -> None:
    payload = []
    for record in records:
        payload.append(
            {
                "id": record["id"],
                "title": record["title"],
                "family": record["familyLabel"],
                "familyKey": record["familyKey"],
                "price": record["price"],
                "delivery": record["delivery"],
                "buyer": record["buyer"],
                "image": f"assets/{record['image']}",
                "detailUrl": f"itens/{record['slug']}/",
                "checkoutUrl": record.get("checkoutUrl"),
            }
        )
    js = "window.SOS_LOW_TICKET_PRODUCTS = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n"
    (WEB / "catalog-data.js").write_text(js, encoding="utf-8")


def write_operations_docs(records: list[dict[str, str]]) -> None:
    """Write the human-facing pricing and release index used by the operator.

    Prices are deliberately labelled as tests: the document gives Francisco a
    launch hypothesis without pretending that fees, refund rate or conversion
    have already been validated in Cakto.
    """
    by_family: dict[str, list[dict[str, str]]] = {}
    for record in records:
        by_family.setdefault(record["familyLabel"], []).append(record)

    pricing_lines = [
        "# Precificação e plano de publicação — Low-ticket SOS Vendas",
        "",
        f"> Atualizado em {date.today().isoformat()} · 48 produtos preparados para validação na Cakto.",
        "",
        "## Regra de preço",
        "Os valores abaixo são preço-teste de pagamento único. Eles compram um artefato aplicável — não implantação, suporte ilimitado, tráfego, consumo de API ou garantia de resultado. Antes de publicar, conferir taxa da Cakto, política de reembolso, e-mail de suporte e o que será efetivamente entregue.",
        "",
        "## Mapa dos produtos",
        "| ID | Produto | Família | Preço-teste | Para quem |",
        "|---:|---|---|---:|---|",
    ]
    for record in records:
        pricing_lines.append(
            f"| {record['id']} | {record['title']} | {record['familyLabel']} | R$ {record['price']},00 | {record['buyer']} |"
        )

    pricing_lines.extend(["", "## Pacotes por trabalho", "Os pacotes abaixo são hipóteses para aumentar ticket sem misturar personas. Publicar depois de validar os itens avulsos.", ""])
    bundles = [
        ("WhatsApp Essencial", "01–06", records[0:6], 97),
        ("Apresentação que vende", "07–12", records[6:12], 147),
        ("Conteúdo que inicia conversas", "13–18", records[12:18], 167),
        ("Agenda sem retrabalho", "19–24", records[18:24], 127),
        ("Relacionamento que continua", "25–30", records[24:30], 167),
        ("Ferramentas de decisão comercial", "31–36", records[30:36], 197),
        ("Time com padrão", "37–42", records[36:42], 197),
        ("Implantação de IA profissional", "43–48", records[42:48], 247),
    ]
    pricing_lines.extend(["| Pacote | Itens | Soma avulsa | Preço-teste do pacote |", "|---|---:|---:|---:|"])
    for name, ids, items, bundle_price in bundles:
        standalone = sum(item["price"] for item in items)
        pricing_lines.append(f"| {name} | {ids} | R$ {standalone},00 | R$ {bundle_price},00 |")

    pricing_lines.extend(
        [
            "",
            "## Ordem de validação",
            "1. Conferir os oito primeiros produtos de cada família e garantir que a entrega da Cakto corresponde ao ZIP.",
            "2. Publicar os 48 avulsos com checkout próprio quando a sessão/API estiver disponível.",
            "3. Medir cliques, compras, reembolsos e pedidos de suporte por família por 14 dias.",
            "4. Só então ativar os oito pacotes e ajustar preço pela evidência.",
            "",
            "## Estado de publicação",
            "A vitrine está ativa com páginas individuais e CTA seguro. O manifesto `cakto-import.json` mantém `checkoutUrl: null` até a criação e conferência da oferta autenticada; não usar URL inventada.",
            "",
        ]
    )
    (OUT / "PRECIFICACAO.md").write_text("\n".join(pricing_lines), encoding="utf-8")

    root_lines = [
        "# Low-ticket SOS Vendas",
        "",
        "Coleção de 48 produtos digitais de aplicação única para WhatsApp, operação comercial e implantação profissional de IA.",
        "",
        "## O que está neste diretório",
        "- `pXX-*/`: pacote individual com `material.pdf`, fonte editável, página de venda e README.",
        "- `releases/`: ZIP pronto para anexar ou entregar.",
        "- `cakto-import.json`: manifesto de importação com preço, imagem, página, entrega e estado do checkout.",
        "- `cakto-checkouts.json`: fatos confirmados no painel Cakto que sobrevivem à regeneração da vitrine.",
        "- `landing/produtos/downloads/`: ZIPs usados como entrega por e-mail da Cakto até a migração para uma área de membros.",
        "- `PRECIFICACAO.md`: mapa de preços-teste e hipóteses de pacote.",
        "",
        "## Regenerar",
        "```bash",
        "python3 scripts/low-ticket/generate_products.py",
        "```",
        "",
        "## Publicação segura",
        "A vitrine pública já lista os 48 itens. A publicação na Cakto exige sessão autenticada e conferência do conteúdo entregue; por isso os checkouts permanecem nulos no manifesto até a etapa de importação. Os links em `downloads/` são provisórios e compartilháveis; migrar para Cakto Members antes de escalar anúncios.",
        "",
    ]
    (OUT / "README.md").write_text("\n".join(root_lines), encoding="utf-8")


def write_detail_page(record: dict[str, str]) -> None:
    page_dir = WEB / "itens" / record["slug"]
    page_dir.mkdir(parents=True, exist_ok=True)
    image_url = f"../../assets/{record['image']}"
    checkout_url = record.get("checkoutUrl")
    cta_href = checkout_url or f"https://wa.me/5549988447562?text=Ol%C3%A1!%20Quero%20receber%20o%20checkout%20de%20{record['title'].replace(' ', '%20')}."
    cta_label = "Comprar agora" if checkout_url else "Quero receber o material"
    cta_note = "Checkout ativo na Cakto." if checkout_url else "O checkout será ativado após a conferência final da oferta."
    body = f"""<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <meta name="theme-color" content="#0b0f17" />
  <title>{escape(record['title'])} — SOS Vendas</title>
  <meta name="description" content="{escape(record['promise'].capitalize())}." />
  <link rel="canonical" href="https://iaparavendas.tech/produtos/itens/{record['slug']}/" />
  <link rel="icon" type="image/svg+xml" href="../../../assets/favicon.svg" />
  <link rel="stylesheet" href="../../styles.css?v=20260905-1500" />
  <meta property="og:image" content="https://iaparavendas.tech/produtos/{image_url[6:]}" />
</head>
<body class="product-detail-page">
  <header class="catalog-header">
    <a class="brand" href="../../" aria-label="SOS Vendas, vitrine"><img src="../../../assets/logo.svg" alt="SOS Vendas" /></a>
    <nav aria-label="Navegação do produto"><a href="../../">Todos os produtos</a><a href="../../../eko/">EKO</a></nav>
    <a class="catalog-header-cta" href="../../#catalogo-low-ticket">Ver catálogo <span aria-hidden="true">↓</span></a>
  </header>
  <main>
    <section class="detail-hero">
      <div class="catalog-wrap detail-grid">
        <div class="detail-copy">
          <p class="eyebrow"><i aria-hidden="true"></i>{escape(record['familyLabel'])}</p>
          <h1>{escape(record['title'])}</h1>
          <p class="detail-lead">{escape(record['promise'].capitalize())}.</p>
          <p class="detail-buyer"><strong>Para:</strong> {escape(record['buyer'])}.</p>
          <div class="detail-price"><strong>R$ {record['price']}</strong><span>pagamento único · preço-teste</span></div>
          <a class="catalog-button" data-product-checkout href="{escape(cta_href)}" target="_blank" rel="noreferrer">{cta_label} <span aria-hidden="true">↗</span></a>
          <p class="detail-note">{cta_note}</p>
        </div>
        <figure class="detail-visual"><img src="{image_url}" alt="Imagem editorial da família {escape(record['familyLabel'])}" width="1254" height="1254" /></figure>
      </div>
    </section>
    <section class="detail-section">
      <div class="catalog-wrap detail-content">
        <p class="eyebrow">Entrega</p>
        <h2>Um artefato para usar na próxima ação.</h2>
        <p>{escape(record['delivery']).capitalize()}.</p>
        <div class="detail-delivery-grid">
          <div><span>01</span><strong>Material editável</strong><p>Campos claros para adaptar ao processo real.</p></div>
          <div><span>02</span><strong>Exemplo fictício</strong><p>Uma referência preenchida sem dados de clientes.</p></div>
          <div><span>03</span><strong>Checklist</strong><p>Critérios para revisar antes de usar ou publicar.</p></div>
        </div>
      </div>
    </section>
    <section class="detail-section detail-section-soft">
      <div class="catalog-wrap detail-content">
        <p class="eyebrow">Uso responsável</p>
        <h2>Clareza para adaptar. Limite para não prometer.</h2>
        <p>Troque campos entre colchetes, use informações confirmadas e encaminhe exceções para uma pessoa responsável. O material não inclui implantação, operação ilimitada ou garantia de resultado.</p>
        <a class="catalog-text-link" href="../../#catalogo-low-ticket">Conhecer outras soluções <span aria-hidden="true">→</span></a>
      </div>
    </section>
  </main>
  <footer class="catalog-footer"><div class="catalog-wrap catalog-footer-inner"><a class="brand" href="../../"><img src="../../../assets/logo.svg" alt="SOS Vendas" /></a><p>Produtos e serviços para continuidade comercial no WhatsApp.</p><small>© 2026 MCT LTDA · Chapecó, SC</small></div></footer>
  <script src="../../../app.js?v=20260905-1500"></script>
</body>
</html>
"""
    (page_dir / "index.html").write_text(body, encoding="utf-8")


def load_cakto_checkouts() -> dict[str, dict[str, str]]:
    """Load only operator-recorded Cakto facts; never invent checkout URLs."""
    if not CAKTO_CHECKOUTS_PATH.exists():
        return {}
    payload = json.loads(CAKTO_CHECKOUTS_PATH.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise RuntimeError(f"Expected an object in {CAKTO_CHECKOUTS_PATH}")
    return {str(key).zfill(2): value for key, value in payload.items() if isinstance(value, dict)}


def main() -> None:
    records = parse_catalog()
    cakto_checkouts = load_cakto_checkouts()
    OUT.mkdir(parents=True, exist_ok=True)
    RELEASES.mkdir(parents=True, exist_ok=True)
    for record in records:
        product_dir = OUT / record["slug"]
        product_dir.mkdir(parents=True, exist_ok=True)
        material = material_markdown(record)
        sales_page = sales_page_markdown(record)
        tool_file = build_tool_file(record, product_dir)
        if tool_file:
            record["toolFile"] = tool_file
        readme = readme_markdown(record)
        (product_dir / "material.md").write_text(material, encoding="utf-8")
        (product_dir / "pagina-venda.md").write_text(sales_page, encoding="utf-8")
        (product_dir / "README.md").write_text(readme, encoding="utf-8")
        pdf_path = product_dir / "material.pdf"
        build_pdf(record, material, pdf_path)
        image_dst = product_dir / "thumbnail.png"
        if image_dst.exists():
            image_dst.unlink()
        archive_path = RELEASES / f"{record['slug']}.zip"
        with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            package_files = ("README.md", "material.md", "material.pdf", "pagina-venda.md")
            if tool_file:
                package_files += (tool_file,)
            for filename in package_files:
                archive.write(product_dir / filename, arcname=filename)
        record["release"] = str(archive_path.relative_to(ROOT))
        record["salesPage"] = f"https://iaparavendas.tech/produtos/itens/{record['slug']}/"
        record["imageUrl"] = f"https://iaparavendas.tech/produtos/assets/{record['image']}"
        record["downloadUrl"] = f"https://iaparavendas.tech/produtos/downloads/{record['slug']}.zip"
        cakto_fact = cakto_checkouts.get(record["id"], {})
        record["caktoProductId"] = cakto_fact.get("productId")
        record["checkoutUrl"] = cakto_fact.get("checkoutUrl")
        record["publicationStatus"] = cakto_fact.get("publicationStatus", "prepared_pending_cakto_access")
        DOWNLOADS.mkdir(parents=True, exist_ok=True)
        shutil.copy2(archive_path, DOWNLOADS / archive_path.name)
        write_detail_page(record)
    write_catalog_js(records)
    write_operations_docs(records)
    manifest = {
        "generatedAt": date.today().isoformat(),
        "source": "docs/CATALOGO_POSSIBILIDADES_LOW_TICKET_2026-09-05.md",
        "note": "Checkout URLs are populated only from products/low-ticket/cakto-checkouts.json after visual verification in Cakto.",
        "products": records,
    }
    (OUT / "cakto-import.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    summary = {
        "products": len(records),
        "pdfs": len(list(RELEASES.glob("*.zip"))),
        "detailPages": len(list((WEB / "itens").glob("*/index.html"))),
    }
    print(json.dumps(summary, ensure_ascii=False))


if __name__ == "__main__":
    main()
